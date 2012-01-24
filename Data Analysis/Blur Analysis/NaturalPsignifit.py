#Using Psignifit to bootstrap the natural scenes data - Dec 2011

from psychopy import data, gui, misc, core
import pypsignifit as psi
import numpy as np
import matplotlib, os
from scipy import stats as stats
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.reader.excel import load_workbook
from openpyxl.cell import get_column_letter


analysisType = 'byCond' #'byPicture

files = gui.fileOpenDlg('.')
if not files:
    core.quit()
   
#Divide the data by picture/condition
sortedFiles = {}

if analysisType == 'byCond':
    sortedFiles['ChromFiles'], sortedFiles['AchromFiles'], sortedFiles['LumFiles'], sortedFiles['IsolumFiles'] = [], [], [], []
    for thisFileName in files:
        thisDat = misc.fromFile(thisFileName)
        if thisDat.extraInfo['Chromatic Blur']=='y':
            if thisDat.extraInfo['Luminance']==1:
                (sortedFiles['ChromFiles']).append(thisFileName)
            if thisDat.extraInfo['Luminance']==0:
                (sortedFiles['IsolumFiles']).append(thisFileName)
        if thisDat.extraInfo['Luminance Blur']=='y':
            if thisDat.extraInfo['Chromaticity']==1:
                (sortedFiles['LumFiles']).append(thisFileName)
            if thisDat.extraInfo['Chromaticity']==0:
                (sortedFiles['AchromFiles']).append(thisFileName)
                
if analysisType == 'byPicture':
    sortedFiles['LeafFiles'], sortedFiles['PansyFiles'], sortedFiles['PelicanFiles'], sortedFiles['PumpkinFiles'] = [], [], [], []
#To divide by image
    pictDict = thisDat.extraInfo['images']
    pic = pictDict.keys()

    if pic[0]== 'Leaf512.jpg':
        thisFileName.append(LeafFiles)
    if pic[0] == 'Pansy512.jpg':
        thisFileName.append(PansyFiles)
    if pic[0] == 'Pelican512.jpg':
        thisFileName.append(PelicanFiles)
    if pic[0] =='Pumpkin512.jpg':
        thisFileName.append(PumpkinFiles)
    analysisType = 'byPicture'
#Get all the data from all the files

for thisGroup, thisFile in sortedFiles.iteritems():
    allIntensities, allResponses, extraInfo, position, reversalIntensities = [],[],[],[],[]

    for thisFileName in thisFile:
        thisDat = misc.fromFile(thisFileName)
        assert isinstance(thisDat, data.StairHandler)
        allIntensities.append(thisDat.intensities)
        allResponses.append(thisDat.data)
        extraInfo.append(thisDat.extraInfo)
        reversalIntensities.append(thisDat.reversalIntensities)
        
    #fetch all the details about the trial
        participant = thisDat.extraInfo['participant']
        if thisDat.extraInfo['Chromatic Blur']=='y':
            if thisDat.extraInfo['Luminance']==1:
                condition = 'ChromBlur'
            if thisDat.extraInfo['Luminance']==0:
                condition = 'Isolum'
        if thisDat.extraInfo['Luminance Blur']=='y':
            if thisDat.extraInfo['Chromaticity']==1:
                condition = 'LumBlur'
            if thisDat.extraInfo['Chromaticity']==0:
                condition = 'Achrom'
                
        pictDict = thisDat.extraInfo['images']
        pic = pictDict.keys()

        if pic[0]== 'Leaf512.jpg':
            picture = 'Leaf'
        if pic[0] == 'Pansy512.jpg':
            picture = 'Pansy'
        if pic[0] == 'Pelican512.jpg':
            picture = 'Pelican'
        if pic[0] =='Pumpkin512.jpg':
            picture = 'Pumpkin'
            
    #reorganise data for psignifit
    newIntensities = []
    for n in range(len(allIntensities)):
        allIntensities[n]=np.array(allIntensities[n])
        reversalIntensities[n]=np.array(reversalIntensities[n])
        
    combinedInten, combinedResp, combinedN = \
                 data.functionFromStaircase(allIntensities, allResponses, bins = 'unique')

    expData = np.c_[combinedInten, combinedResp, combinedN]

    nafc = 1

    #Not sure if these settings are right
    constraints = ('unconstrained', 'unconstrained', 'Beta(2,20)')

    boots = psi.BootstrapInference(expData, core='ab', sigmoid='gauss', priors=constraints, nafc=2)

    boots.sample(5000)

    psi.plotSensitivity(boots)

    #analysis
    thresh = boots.getThres(0.75)

    print 'thresh', thresh
    print 'slope', boots.getSlope()
    #print 'jnd', (boots.getThres(0.75)-boots.getThres(0.25))

    estimates = boots.mcestimates
    betas = []

    for n in range(len(estimates)):
        temp = estimates[n]
        betas.append(temp[1])

    meanBeta = np.mean(betas)
    print 'mean', meanBeta, 'median', np.median(betas), 'mode', stats.mode(betas)
    SE = (np.std(betas))/(np.sqrt(len(betas)))

    upperBetaCI = meanBeta + (1.96 * SE)
    lowerBetaCI = meanBeta - (1.96 * SE)

    print 'upper', upperBetaCI, 'lower', lowerBetaCI
    print 'std BETA', np.std(betas)

    print 'alpha, beta, lapse', boots.estimate

    threshUpper = boots.getCI(0.75)[0]
    threshLower = boots.getCI(0.75)[1]

    print 'thresh ci', boots.getCI(0.75)
    print 'slope ci', boots.getCI(0.75, thres_or_slope = 'slope')

    #set up Excel file
    stimOut = {'thresh' : 1, 'thresh ci upper' : 2, 'thresh ci lower' : 3}

#    fileName = ('LumReverse_%s.xlsx') %(participant)
    fileName = 'RJSLumChromReversal.xlsx'

    if os.path.isfile(fileName):
        wb = load_workbook(fileName)
        newWorkbook = False
    else:
        wb = Workbook()
        newWorkbook = True

    ew = ExcelWriter(workbook = wb)

    if analysisType == 'byCond':
        sheetName = '%s' %(condition)
    else:
        sheetName = '%s%s' %(condition, picture)

    if newWorkbook:
        ws = wb.worksheets[0]
        ws.title=sheetName
    else:
        ws=wb.create_sheet()
        ws.title=sheetName

    def _getExcelCellName(col, row):
        """Returns the excel cell name for a row and column (zero-indexed)

        >>> _getExcelCellName(0,0)
        'A1'
        >>> _getExcelCellName(2,1)
        'C2'
        """
        return "%s%i" %(get_column_letter(col+1), row+1)

    #enter data into Excel sheet
    for colN, heading in enumerate(stimOut):
        print heading
        ws.cell(_getExcelCellName(col=colN, row=0)).value=unicode(heading)
        
    ws.cell('A2').value = unicode(thresh)
    ws.cell('B2').value = unicode(threshUpper)
    ws.cell('C2').value = unicode(threshLower)

    ew.save(filename = fileName)
   
print 'TA DA!!'