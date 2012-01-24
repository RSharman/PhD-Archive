# 2D Fast Fourier Transform on the lum, lm and s components of natural scenes - March 2011

#IF YOU DO NOT CHANGE THE DESTINATION FILE NAME, IT WILL OVER WRITE ANY
#EXISTING FILE WITH THE SAME NAME

from psychopy import misc, core, visual, event, data, gui
from scipy import fftpack
import scipy, Image, copy, pylab, glob
import numpy as np
import radialProfile, radial_data
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
#
#files = glob.glob('C:\Documents and Settings\lpxrs\My Documents\Colour Data\Image Statistics\McGill Image Library/Textures/*.tif')
#files = gui.fileOpenDlg('.', allowed = "JPEG files (*.jpg) | *.jpg")
files = gui.fileOpenDlg('.', allowed = "TIFF files (*.tif) | *.tif")
if not files:
    core.quit()

counter = 2

#Name of file where data will be saved
#fName = "Textures2.xlsx"

#Create Excel File
wb = Workbook()
ew = ExcelWriter(workbook = wb)

ws = wb.get_active_sheet()
ws.title = "Raw Data"

#Analyse each file in the selection
for thisFileName in files:

    OrigImage = thisFileName

    OrigPicture = Image.open(OrigImage).transpose(Image.FLIP_TOP_BOTTOM)
    sizer = np.array(OrigPicture)

#Measure the size of the image in order to crop out the centre 512x512 block
#    print sizer.shape

    if len(sizer[0,:,:]>512):
        xremove = (len(sizer[0,:,:])-512)/2
    if len(sizer[:,0,:]>512):
        yremove = (len(sizer[:,0,:])-512)/2

    #Crop to centre 512x512
    box = (xremove, yremove, 512+xremove, 512+yremove)
    picture = np.array(OrigPicture.crop(box)) /127.5-1

#Convert to DKL and separate into channels
    dklPicture = misc.rgb2dklCart(picture, conversionMatrix=None)
    dklPicture = np.array(dklPicture)

    lum = copy.copy(dklPicture[:,:,0])
    lm = copy.copy(dklPicture[:,:,1])
    s = copy.copy(dklPicture[:,:,2])

#Run 2d Fast Fourier Transform on each Channel
    fftLum = fftpack.fft2(lum)
    fftLm = fftpack.fft2(lm)
    fftS = fftpack.fft2(s)

    F2Lum = fftpack.fftshift(fftLum)
    F2Lm = fftpack.fftshift(fftLm)
    F2S = fftpack.fftshift(fftS)

    #Not sure why this is squared www.astrobetter.com/fourier-transforms-of-images-in-python/
    #Or why a log is take before plotting, from the same source
    F3Lum = abs(F2Lum)**2
    F3Lm = abs(F2Lm)**2
    F3S = abs(F2S)**2

    #Plotting kspace
#    pylab.figure(1)
#    pylab.clf()
#    pylab.imshow(np.log10(F3Lum))

    #Calculating the azimuthally average 1D power spectrum using radialProfile
    A1Lum = radialProfile.azimuthalAverage(F3Lum)
    A1Lm = radialProfile.azimuthalAverage(F3Lm)
    A1S = radialProfile.azimuthalAverage(F3S)

#Plot radial Profile
#    pylab.figure(2)
#    pylab.clf()
#    pylab.semilogy(A1Lum, label = 'Lum')
#    pylab.semilogy(A1Lm, label = 'Lm')
#    pylab.semilogy(A1S, label = 'S')
#    pylab.title(OrigImage)
#    pylab.legend()
#    pylab.xlim(0, 256)
#    pylab.xlabel('Spatial Frequency')
#    pylab.ylabel('Power Spectrum')
#    
    pylab.figure(2)
    pylab.clf()
    
    pylab.loglog(A1Lum, label = 'lum')
    pylab.loglog(A1Lm, label = 'lm')
    pylab.loglog(A1S, label = 's')
    pylab.legend()
    
    pylab.show()

    #Take mean of the last 128 points
    LumMean = np.mean(A1Lum[127:256])
    LmMean = np.mean(A1Lm[127:256])
    SMean = np.mean(A1S[127:256])
    print LumMean
    print LmMean
    print SMean

    #Save Data to excel
#    lumCell = 'A'+str(counter)
#    lmCell = 'B'+str(counter)
#    sCell = 'C'+str(counter)
#     
#     
#    ws.cell(lumCell).value = LumMean
#    ws.cell(lmCell).value = LmMean
#    ws.cell(sCell).value = SMean
#    
#    counter +=1
#
#ws.cell('A1').value = 'Lum'
#ws.cell('B1').value = 'Lm'
#ws.cell('C1').value = 'S'
#ew.save(filename = fName)
#
#print 'Your data has been saved to %s' %fName




